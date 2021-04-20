#!/usr/bin/env python
# coding: utf-8

# In[1]:


import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
get_ipython().run_line_magic('matplotlib', 'inline')
sns.set()

from matplotlib import *
f = 'C:/Windows/Fonts/malgun.ttf'
rc('font', family=font_manager.FontProperties(fname=f).get_name())
rcParams['axes.unicode_minus'] = False


# ### 엑셀 시트 추가

# In[2]:


from openpyxl import load_workbook
def excel_add_sheet(filename, df, sheet_name):
    writer = pd.ExcelWriter(filename, 'openpyxl')
    try:
        writer.book = load_workbook(filename)   # try to open an existing workbook

        if sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name) # 삭제할 sheet 위치
            writer.book.remove(writer.book.worksheets[idx]) # 삭제 
            writer.book.create_sheet(sheet_name, idx) # 생성

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    df.to_excel(writer, sheet_name, index=False)
    writer.save()
    
# df = pd.DataFrame([1])
# excel_add_sheet('test.xlsx', df, 'testtt')


# ### (함수) 시가총액 정리

# In[3]:


def get_mkt_price(filepath):
    mkt = pd.read_excel(filepath, encoding='cp949', header=[0])
    mkt['종목코드'] = mkt['종목코드'].astype('str').str.zfill(6) #코드유지
    mkt['시가총액'] = mkt['시가총액'].str.replace(' 억',"")
    mkt['시가총액'] = mkt['시가총액'].str.replace('\,',"").astype('int')
    mkt['현재가'] = mkt['현재가'].str.replace('\,',"").astype('int')

    return mkt

# get_mkt_price('data/시총.xlsx').head(2)


# ### (함수) 거래내역 정리

# In[4]:


def get_deal_history(filepath):
    history = pd.read_excel(filepath, encoding='cp949', header=[0,1])
    
    # shift
    history[('이자','신용이자')] = history[('수량','단가')].shift(-1)
    history[('수수료','제세금')] = history[('종목코드','종목명')].shift(-1)
    history[('융자/대주금액','미수발생/변제금')] = history[('거래일자','취소여부')].shift(-1)
    history[('잔량','RP+예수금')] = history[('거래NO','원거래NO')].shift(-1)
    
    # 필드정리 후 필요정보만 사용 + index정리 
    history.columns = ['거래일자','거래NO','거래적요','종목코드','수량','단가','종목명','거래금액','취소여부','원거래NO']
    history = history[history['거래적요'].notnull()]
    history = history[history['수량'] > 0]
    history = history.reset_index(drop=True)

    # 취소거래 & 원거래 제외
    cancel_idx = history['거래NO'][history['취소여부'] == 'Y'].tolist()
    src_idx = history['원거래NO'][history['취소여부'] == 'Y'].drop_duplicates().astype('int').tolist()
    history = history[~history['거래NO'].isin(cancel_idx + src_idx)]
    
    # 컬럼 전처리
    history['거래적요'] = history['거래적요'].map({'주식매수입고':'매수', '주식매도출고':'매도'})
    history['거래일자'] = pd.to_datetime(history['거래일자'], format='%Y/%m/%d')
    history['종목코드'] = history['종목코드'].str.replace('A','')
    history['단가'] = history['단가'].str.replace(',','').astype('int')
    history['거래금액'] = history['수량'] * history['단가']
    history = history[['거래일자','거래적요','종목코드','수량','단가','종목명','거래금액']]
    
    return history

# get_deal_history('data/거래.xls').head(2)


# ### (함수) 재무수치

# In[5]:


def get_finance(filepath):
    finance = pd.read_excel(filepath, sheet_name='제조', encoding='cp949', header=[1,2]) # '제조' 또는 '제조(연결)'
#     finance = pd.read_pickle('pkl/finance_raw.pkl')
    
    # 멀티컬럼 정리
    lv0 =  finance.columns.levels[0].str.replace('\n','')
    finance.columns = finance.columns.set_levels(lv0, level=0)
    finance.columns = [ str(lv0)+'_'+str(lv1) for lv0, lv1 in  finance.columns ]
    
    # 재무지표 기준연월
    yyyymm = filepath.split('.')[0][-6:]
    
    # 데이터 필터 : 코스피만
    finance = finance[finance['시장_Unnamed: 0_level_1'] == 'KS']
    
    # 컬럼 전처리
    AstDpt_col = ['종목코드_Unnamed: 1_level_1', '회사명_Unnamed: 2_level_1', '자산총계_'+yyyymm, '부채총계_'+yyyymm]
    finance_col = finance.columns[finance.columns.str.contains('영업이익\(보고서')].tolist()
    finance = finance[AstDpt_col + finance_col]
    finance = finance.reset_index(drop=True)
    
    finance.columns = finance.columns.str.replace('영업이익\(보고서기재\)_','')
    finance.columns = finance.columns.str.replace('/\누적','')
    finance.columns = finance.columns.str.replace('_Unnamed: 1_level_1','')
    finance.columns = finance.columns.str.replace('_Unnamed: 2_level_1','')
    finance = finance[finance.columns[~finance.columns.str.contains('3개월|비교')]]
    finance.columns = finance.columns.str.replace('_'+yyyymm,'')
    finance = finance.rename(columns={'회사명':'종목명'})

    finance['종목코드'] = finance['종목코드'].str.replace('A','')
    finance.iloc[:,2:] = finance.iloc[:,2:].applymap(lambda x : round(x/100000,1)) #억 단위로 
    
    return finance

# get_finance('data/재무_202006.xlsx').head(2)


# ### (함수) 잠정실적 정보

# In[6]:


def get_pre_result(filepath):
    pre = pd.read_excel(filepath, encoding='cp949', header=[0,1])
    
    # 컬럼 전처리
    pre = pre[[ ('종목코드','Unnamed: 0_level_1'), ('연결','구분'),( '분기실적(억원)', '영업이익' ) ]]
    pre.columns = ['종목코드','연결별도','다음Q']
    pre = pre[pre['연결별도'] == '별도']

    pre['다음Q'] = (pre['다음Q'].fillna(0)).astype('int')
    pre['종목코드'] = pre['종목코드'].str.replace('A','')
    
    # 데이터 필터 : 잠정실적 있는 종목만
    pre = pre[pre['다음Q']>0]
    pre = pre[['종목코드','다음Q']]
    
    return pre

# get_pre_result('data/잠정.xlsx').head(2)


# ### (함수) 최근 1년이익 계산

# In[7]:


def get_income(finance, pre, yyyymm): 
    # 재무자료 기준연월
    base = pd.datetime(int(yyyymm[:4]), int(yyyymm[-2:]),15)    
    base_Q = base.strftime('%Y%m')
    b1y_Q = (base - pd.Timedelta(1, unit='Y')).strftime('%Y%m')
    b1y_4Q = (base - pd.Timedelta(1, unit='Y')).strftime('%Y') + "12"  
    
    base2_Q = base + pd.Timedelta(3, unit='M') # 잠정실적용 변수 : 기준점 + 3개월
    base2_b1y_Q = (base2_Q - pd.Timedelta(1, unit='Y')).strftime('%Y%m')
    
    # 최근 1년 영업이익 전체 구하기
    if base_Q[-2:] == '12':
        finance['최근4Q'] = finance[base_Q]
    else:
        finance['최근4Q'] = finance[base_Q] + ( finance[b1y_4Q] - finance[b1y_Q] ) 
        
    finance = finance[finance['최근4Q'].notnull()]
    
    # 잠정실적 결합 + 잠정실적 나온 종목은 1년 영업이익값 조정
    income = pd.merge(finance, pre, how='left', on='종목코드')
    
    if base_Q[-2:] == '12':
        income.loc[income['다음Q'].notnull(), '최근4Q'] = income['최근4Q'] + income['다음Q'] - income[base2_b1y_Q]
    else:
        income.loc[income['다음Q'].notnull(), '최근4Q'] = income['최근4Q'] + income['다음Q'] - ( income[base2_b1y_Q] - income[b1y_Q] ) 
        
    return income


# In[8]:


# finance = get_finance('data/재무_202006.xlsx')
# finance.to_pickle('pkl/finance.pkl')
# pre = get_pre_result('data/잠정.xlsx')
# get_income(finance,pre,'202006').head(2)


# ### (함수)마법공식 (미보유)topn 종목 받기

# In[16]:


def get_stock_topn(topn, yyyymm, min_mkt_price, none_port): # none_port - 1: 미보유 중, 0:전체 중
    # 잠정실적 반영
    finance = pd.read_pickle('pkl/finance.pkl')
    pre = get_pre_result('data/잠정.xlsx')
    income = get_income(finance, pre, yyyymm)
    
    # 시총정보 추가
    mkt = get_mkt_price('시총.xlsx')
    fin_rank = pd.merge( income[['종목코드','종목명','자산총계','부채총계','최근4Q']], mkt, how='left', on=['종목코드','종목명'])
    
    # 시총제한  
    fin_rank = fin_rank[fin_rank['시가총액'] > min_mkt_price]
    
    # 자본수익률 : 영업이익 / 자산총계 
    # 이익수익률 : 영업이익 / ( 시가총액 + 부채총계 )    # EV: 기업가치 = 시총+부채
    fin_rank['자본수익률'] = (fin_rank['최근4Q'] / fin_rank['자산총계'] ).round(2)
    fin_rank['이익수익률'] = (fin_rank['최근4Q'] / ( fin_rank['시가총액'] + fin_rank['부채총계'] )).round(2)

    # 재무지표 존재 모든 종목들의 순위 매겨봄
    fin_rank['자본수익률순위'] = fin_rank['자본수익률'].rank(ascending=False, method='min')
    fin_rank['이익수익률순위'] = fin_rank['이익수익률'].rank(ascending=False, method='min')
    fin_rank['순위합산'] = fin_rank['자본수익률순위'] + fin_rank['이익수익률순위']
    fin_rank = fin_rank.sort_values('순위합산')
    fin_rank['최종순위'] =  fin_rank['순위합산'].rank(ascending=True, method='min')
    fin_rank = fin_rank.reset_index(drop=True)
    
    # 현 포트에 매수되어 있는지 정보 추각
    port_yn = pd.read_excel('data/포트.xls')[['종목코드','매수금액']]
    port_yn['종목코드'] = port_yn['종목코드'].str.replace('A','')
    fin_rank = pd.merge(fin_rank, port_yn, how='left', on='종목코드').fillna("")
        
    if none_port == 1 :   # 미보유 중 topn
        return fin_rank[fin_rank['매수금액']==""][:topn]
    else:                            # 전체  중 topn
        return fin_rank[:topn]
    
# get_stock_topn(topn=3, yyyymm='202006', min_mkt_price=500, none_port=1) # none_port - 1: 미보유 중, 0:전체 중


# ### (함수) 포트 정리 ※제외항목list

# In[14]:


def get_port(except_list, yyyymm, rank_mkt_price):
    port = pd.read_excel('포트.xls', encoding='cp949', header=[0])
    port = port[port['체결잔고'] != 0]
    port = port[~port['종목명'].isin(except_list)]
    
    port['종목코드'] = port['종목코드'].str.replace('A',"")
    port['매수가격'] = (port['매수금액'] / port['체결잔고']).astype('int')
    port = port[['종목코드', '종목명','매수가격', '체결잔고', '매수금액', '평가금액','평가손익','수익률']]
    port['보유여부'] = "보유중" 
    
    # 투자기준월 정보 추가
    base = pd.read_excel('../(일일)성과측정/보유종목(투자시점).xlsx')[['투자기준월','종목코드','매도월']]
    base = base[base['매도월'].isnull()][['투자기준월','종목코드']]
    base['종목코드'] = base['종목코드'].astype('str').str.zfill(6)
    port = pd.merge(port, base, how='left', on='종목코드').fillna("")  
    port = port.sort_values('투자기준월', ascending=False).reset_index(drop=True)
    port = port[['투자기준월','종목코드', '종목명', '매수가격', '체결잔고', '매수금액', '평가금액', '평가손익', '수익률', '보유여부']]

    # 순위정보 포함 여부
    if rank_mkt_price == -1:
        return port
    else :         
        fin_rank = get_stock_topn(topn=2000, yyyymm='202006',  min_mkt_price=500, none_port=0)[['종목코드','최종순위','현재가']]
        port = pd.merge(port, fin_rank, how='left', on='종목코드')  
        port = port.reset_index(drop=True)    
        return port
    
get_port([], '202006', 500).head()


# In[11]:


def shot_today_score(shot, today, add_invest, stock, rp, kospi):
    # 해당 날짜 데이터가 혹시 있으면 해당 위치에 작업진행
    if shot[shot['일자'] == today].shape[0] > 0 :
        idx = shot[shot['일자'] == today].index[0]
    else:
        idx = shot.shape[0]

    shot.loc[ idx, '일자' ] = today
    shot.loc[ idx, '추가투자' ] = add_invest
    shot.loc[ idx, '투자원금' ] = shot.loc[ idx - 1, '투자원금' ] + add_invest

    shot.loc[ idx, '주식평가' ] = stock
    shot.loc[ idx, '예수금' ] = rp
    shot.loc[ idx, '총평가액' ] = stock + rp
    shot.loc[ idx, '손익' ] = shot.loc[ idx, '총평가액' ] - shot.loc[ idx, '투자원금' ]
    shot.loc[ idx, '일변화율' ] = ((shot.loc[ idx, '총평가액' ]-add_invest)/shot.loc[ idx-1, '총평가액' ] - 1) * 100
    shot.loc[ idx, '총변화율' ] = (shot.loc[ idx, '총평가액' ]/shot.loc[ idx    , '투자원금' ] - 1) * 100

    shot.loc[ idx, 'KOSPI' ] = kospi
    shot.loc[ idx, 'K수량' ]   =  shot.loc[ idx-1, 'K수량' ] + int(add_invest / kospi)
    shot.loc[ idx, 'K예수금' ] =  shot.loc[ idx-1, 'K예수금' ] + add_invest - (kospi * int(add_invest / kospi))
    shot.loc[ idx, 'K평가액' ] =  round(shot.loc[ idx, 'KOSPI' ] * shot.loc[ idx, 'K수량' ],0)
    shot.loc[ idx, 'K손익' ] = shot.loc[ idx, 'K평가액' ] - shot.loc[ idx, '투자원금' ]
    shot.loc[ idx, 'K일변화율' ] = ((shot.loc[ idx, 'K평가액' ]-add_invest)/shot.loc[ idx-1, 'K평가액' ] -1) * 100
    shot.loc[ idx, 'K총변화율' ] = (shot.loc[ idx, 'K평가액' ]/shot.loc[ idx    , '투자원금' ] -1) * 100

    shot.loc[ idx, 'BM대비' ] = np.where(shot.loc[ idx, '일변화율'] > shot.loc[ idx, 'K일변화율'], "승",
                                                  ( np.where(shot.loc[ idx, '일변화율'] < shot.loc[ idx, 'K일변화율'], "패", "무")))

    shot.to_excel('shot2.xlsx', index= False)
    return shot


# ### 네이버 수정종가 크롤링

# In[12]:


import requests
from bs4 import BeautifulSoup
import datetime

def get_mPrice_day(코드, nDay):
    today = datetime.datetime.now()
    url = f'https://fchart.stock.naver.com/sise.nhn?symbol={코드}&timeframe=day&count={nDay}&requestType=0'
    
    request_result = requests.get(url)
    bs = BeautifulSoup(request_result.content, 'html.parser')
    chart_data = bs.select('chartdata')
    fetch_item = bs.select('item')
    
    # chartdata형태 : <chartdata count="744" name="삼성전자" origintime="19900103" precision="0" symbol="005930" timeframe="day">
    name = chart_data[0].attrs['name']
    code = chart_data[0].attrs['symbol']
    origintime = chart_data[0].attrs['origintime']
    
    result_dic={}
    li_date = []
    li_mPrice = []

    # item별 형태: <item price_data="20170814|45120|45400|44720|45000|383712"></item>
    for i in range(len(fetch_item)):
        price_data = str(fetch_item[i]).split("\"")[1].split('|') # 파싱
        li_date.append(pd.to_datetime(price_data[0])) # 일자 리스트추가
        li_mPrice.append(float(price_data[4])) # 수정종가 리스트추가
        
    df = pd.DataFrame(li_mPrice, index=li_date, columns = [code])
    return df

# get_mPrice_day('005930', 5) # 일자 ※수정종가 : 액면분할,증자,감자 고려 조정


# ### 포트와 날짜로 주식포트 총 평가액 구해주는 함수

# In[13]:


def get_portAmt(port, date):
    total = 0
    for i, code in enumerate(port['종목코드']):
        total = total + get_mPrice_day(code, 5).loc[date][code] * port['수량'][i] # 일자 ※수정종가 : 액면분할,증자,감자 고려 조정
    return total

